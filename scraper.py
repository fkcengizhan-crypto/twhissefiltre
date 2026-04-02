"""
TradingView BIST Screener - Günlük Veri Çekici
GitHub Actions tarafından her gün otomatik çalıştırılır.
"""

import asyncio
import json
import os
import re
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SCREENER_URL = "https://tr.tradingview.com/screener/"

# Türkçe format için kısaltma eşleşmeleri
def format_hacim(value_str: str) -> str:
    """Ham hacim değerini Türkçe kısaltmalı formata çevirir."""
    if not value_str or value_str in ("—", "-", ""):
        return ""
    
    # Zaten K/M/B ekli gelebilir (TradingView'dan)
    match = re.match(r"^(-?[\d.,]+)\s*([KkMmBb]?)$", value_str.strip())
    if not match:
        return value_str
    
    num_str = match.group(1).replace(",", ".")
    suffix = match.group(2).upper()

    try:
        raw = float(num_str)
    except ValueError:
        return value_str

    multipliers = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
    raw *= multipliers.get(suffix, 1)

    abs_val = abs(raw)
    sign = "-" if raw < 0 else ""

    if abs_val >= 1e9:
        return f"{sign}{abs_val / 1e9:.2f}Mr".replace(".", ",")
    elif abs_val >= 1e6:
        return f"{sign}{abs_val / 1e6:.2f}M".replace(".", ",")
    elif abs_val >= 1e3:
        return f"{sign}{abs_val / 1e3:.2f}B".replace(".", ",")
    else:
        return f"{sign}{int(abs_val)}"


def clean_value(value: str, col_index: int, header: str) -> str:
    """Hücre değerini temizler ve gerekirse dönüştürür."""
    value = value.strip()

    # Fiyat sütunu — " TRY" temizle
    if col_index == 1:
        return value.replace(" TRY", "")

    # Ortalama Hacim sütunu
    if "ort" in header.lower() and "hacim" in header.lower():
        return format_hacim(value)

    # Sayısal sütunlar (index 2-14)
    if 2 <= col_index <= 14:
        value = value.replace("\u2212", "-")  # Unicode eksi işareti
        value = value.replace("%", "")
        value = value.replace(",", ".")
        if value in ("", "—", "-", "\u2014"):
            return ""
    return value


async def scroll_to_load_all(page) -> None:
    """Tüm hisseler yüklenene kadar sayfayı kaydırır."""
    print("🔄 Scroll başlıyor...")

    # Scroll container'ı bul
    container = await page.query_selector(".wrapper-fFDq5D2D")
    if not container:
        container = await page.query_selector("[class*='wrapper-']")

    prev_count = 0
    unchanged = 0
    iteration = 0

    while unchanged < 8 and iteration < 100:
        iteration += 1

        if container:
            await page.evaluate("el => el.scrollTop = el.scrollHeight", container)
        else:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")

        await asyncio.sleep(2)

        current_count = await page.evaluate(
            "document.querySelectorAll('tbody tr').length"
        )

        print(f"  İterasyon {iteration}: {current_count} satır (önceki: {prev_count})")

        if current_count == prev_count:
            unchanged += 1
        else:
            unchanged = 0

        prev_count = current_count

    print(f"✅ Scroll tamamlandı. Toplam {prev_count} satır.")
    await asyncio.sleep(2)


async def extract_data(page) -> tuple[list[str], list[list[str]]]:
    """Tablo başlıklarını ve verilerini çeker."""

    # Başlıklar
    headers = await page.evaluate("""
        () => {
            const cells = document.querySelectorAll('thead th[role="columnheader"], thead th, th');
            return Array.from(cells).map(c => c.textContent.trim()).filter(Boolean);
        }
    """)
    print(f"📋 Başlıklar ({len(headers)} adet): {headers}")

    # Satır verileri
    rows = await page.evaluate("""
        () => {
            const rows = document.querySelectorAll(
                'tbody.tv-data-table__tbody tr.tv-data-table__row.tv-screener-table__result-row, ' +
                'tbody tr.tv-data-table__row, ' +
                'tbody tr[role="row"], ' +
                'tbody tr'
            );
            return Array.from(rows).map(row => {
                const cells = Array.from(row.querySelectorAll('td'));
                return cells.map((cell, i) => {
                    if (i === 0) {
                        const sym = cell.querySelector('.tickerNameBox-ixuo49jq, [class*="tickerName"]');
                        if (sym) return sym.textContent.trim();
                        const a = cell.querySelector('a');
                        if (a) return a.textContent.trim();
                        return cell.textContent.trim().split('\\n')[0].trim();
                    }
                    return cell.textContent.trim();
                });
            }).filter(r => r.length > 0 && r[0]);
        }
    """)

    print(f"✅ {len(rows)} satır çekildi")
    return headers, rows


def build_excel(headers: list[str], rows: list[list[str]], output_path: Path) -> None:
    """Verileri biçimlendirilmiş Excel dosyasına yazar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BIST Screener"

    # Stil tanımları
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4A6CF7")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D0D0D0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill = PatternFill("solid", fgColor="F4F6FF")

    # Başlık satırı
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 30

    # Veri satırları
    for row_idx, row_data in enumerate(rows, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, cell_value in enumerate(row_data):
            header = headers[col_idx] if col_idx < len(headers) else ""
            cleaned = clean_value(cell_value, col_idx, header)

            # Sayıya çevirmeyi dene (metin formatındaki hacim hariç)
            is_hacim_col = "ort" in header.lower() and "hacim" in header.lower()
            numeric = None
            if not is_hacim_col and col_idx > 0:
                try:
                    numeric = float(cleaned)
                except (ValueError, TypeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if is_alt:
                cell.fill = alt_fill

            if numeric is not None:
                cell.value = numeric
                # Yüzde sütunları (index 2-12, RSI ve hacim hariç)
                if 2 <= col_idx <= 12 and "rsi" not in header.lower():
                    cell.number_format = "0.00%"
                    cell.value = numeric / 100
                else:
                    cell.number_format = '#,##0.00'
            else:
                cell.value = cleaned

    # Sütun genişlikleri otomatik ayarla
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for c in row:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

    # Dondur: başlık satırı sabit kalsın
    ws.freeze_panes = "A2"

    # Metadata sayfası
    meta = wb.create_sheet("Bilgi")
    meta["A1"] = "Çekilme Tarihi"
    meta["B1"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    meta["A2"] = "Toplam Hisse"
    meta["B2"] = len(rows)
    meta["A3"] = "Kaynak"
    meta["B3"] = SCREENER_URL

    wb.save(output_path)
    print(f"✅ Excel kaydedildi: {output_path}")


async def main():
    output_dir = Path("data")
    output_dir.mkdir(exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    output_path = output_dir / f"bist_screener_{today}.xlsx"
    latest_path = output_dir / "bist_screener_latest.xlsx"

    # TradingView oturum cookie'si (GitHub Secret'tan okunur)
    tv_session = os.environ.get("TV_SESSION_ID", "")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )

        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1920, "height": 1080},
            locale="tr-TR",
        )

        # Oturum cookie'si varsa ekle
        if tv_session:
            await context.add_cookies([{
                "name": "sessionid",
                "value": tv_session,
                "domain": ".tradingview.com",
                "path": "/",
            }])
            print("✅ TradingView oturumu yüklendi")
        else:
            print("⚠️  TV_SESSION_ID bulunamadı, anonim erişim deneniyor...")

        page = await context.new_page()

        # Bot tespitini zorlaştır
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        """)

        print(f"🌐 Sayfa açılıyor: {SCREENER_URL}")
        await page.goto(SCREENER_URL, wait_until="networkidle", timeout=60_000)
        await asyncio.sleep(3)

        # Tüm verileri yükle
        await scroll_to_load_all(page)

        # Veri çek
        headers, rows = await extract_data(page)

        if not rows:
            raise RuntimeError("Hiç veri çekilemedi! TradingView yapısı değişmiş olabilir.")

        await browser.close()

    # Excel oluştur
    build_excel(headers, rows, output_path)

    # "latest" kopyasını da güncelle
    import shutil
    shutil.copy2(output_path, latest_path)
    print(f"✅ Latest kopyası güncellendi: {latest_path}")

    # Özet
    print(f"\n📊 ÖZET")
    print(f"   Tarih    : {today}")
    print(f"   Hisse    : {len(rows)}")
    print(f"   Sütun    : {len(headers)}")
    print(f"   Dosya    : {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
